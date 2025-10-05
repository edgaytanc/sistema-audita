import openpyxl
from datetime import datetime
from django.db import transaction
from audits.models import Audit
from auditoria.models import BalanceCuentas, RegistroAuxiliar, SaldoInicial
import traceback
import io

# Importadores delegados
from .processors.annual_importer import process_annual_sheet as _process_annual_sheet
from .processors.semestral_importer import process_semestral_sheet as _process_semestral_sheet
from .processors.auxiliary_importer import process_auxiliary_records as _process_auxiliary_records
from .processors.initial_balances_importer import process_initial_balances as _process_initial_balances

class EstadosFinancierosImporter:
    """
    Importador unificado para archivos de estados financieros que contiene
    tanto balances anuales como semestrales en diferentes hojas.
    
    El archivo debe tener:
    - Una hoja "ESTADOS FINANCIEROS ANUAL" con balances anuales (2 fechas)
    - Una hoja "ESTADOS FINANCIEROS SEMESTRALES" con balances semestrales (4 fechas)
    - Opcionalmente, hojas para registros auxiliares y saldos iniciales
    """
    def __init__(self, file_obj, audit_id):
        """
        Inicializa el importador con un archivo en memoria y un ID de auditoría
        
        Args:
            file_obj: Un objeto InMemoryUploadedFile o similar que puede ser leído directamente
            audit_id: ID de la auditoría asociada
        """
        self.file_obj = file_obj
        self.audit_id = audit_id

    def validate_file(self):
        """Valida que el archivo contenga al menos una de las hojas requeridas"""
        try:
            # Leemos directamente desde el objeto de archivo
            file_content = io.BytesIO(self.file_obj.read())
            # Importante: restaurar el puntero para lecturas futuras
            self.file_obj.seek(0)
            
            wb = openpyxl.load_workbook(file_content, read_only=True, data_only=True)
            sheet_names_lower = [name.lower() for name in wb.sheetnames]
            
            # Verificar si existe al menos una de las hojas principales
            if not any('anual' in name or 'semestral' in name for name in sheet_names_lower):
                return False
                
            return True
        except Exception as e:
            return False

    @transaction.atomic
    def process_file(self):
        """Procesa todas las hojas relevantes del archivo"""
        try:
            # Leemos directamente desde el objeto de archivo
            file_content = io.BytesIO(self.file_obj.read())
            # Importante: restaurar el puntero para futuras operaciones
            self.file_obj.seek(0)
            
            wb = openpyxl.load_workbook(file_content, data_only=True)
            
            # Convertir nombres de hojas a minúsculas para comparación
            sheet_names_lower = {name.lower(): name for name in wb.sheetnames}
            
            # Procesamiento de hojas de estados financieros
            anual_processed = False
            semestral_processed = False
            
            # Procesar la hoja de estados financieros anuales si existe
            anual_sheet_name = next((sheet_names_lower[name] for name in sheet_names_lower 
                                    if 'anual' in name and 'semestral' not in name), None)
            if anual_sheet_name:
                self.process_annual_sheet(wb[anual_sheet_name])
                anual_processed = True
            
            # Procesar la hoja de estados financieros semestrales si existe
            semestral_sheet_name = next((sheet_names_lower[name] for name in sheet_names_lower 
                                        if 'semestral' in name), None)
            if semestral_sheet_name:
                self.process_semestral_sheet(wb[semestral_sheet_name])
                semestral_processed = True
            
            # Procesar la hoja de registros auxiliares si existe
            aux_sheet_name = next((sheet_names_lower[name] for name in sheet_names_lower 
                                  if 'auxiliar' in name), None)
            if aux_sheet_name:
                self.process_auxiliary_records(wb[aux_sheet_name])
            
            # Procesar la hoja de saldos iniciales si existe
            initial_sheet_name = next((sheet_names_lower[name] for name in sheet_names_lower 
                                      if 'saldo' in name), None)
            if initial_sheet_name:
                self.process_initial_balances(wb[initial_sheet_name])
            
            # Verificar que se haya procesado al menos una hoja principal
            if not (anual_processed or semestral_processed):
                return False, "❌ No se encontró ninguna hoja válida de estados financieros"
            
            return True, "✅ Importación completada exitosamente"
        except Exception as e:
            error_msg = f"Error procesando archivo: {str(e)}"
            traceback.print_exc()
            return False, error_msg

    # ------------------------------------------------------------------
    #  Nuevos wrappers que delegan al importador especializado
    # ------------------------------------------------------------------

    def process_annual_sheet(self, sheet):
        """Enrutador hacia `imports.annual_importer.process_annual_sheet`."""
        return _process_annual_sheet(sheet, self.audit_id)

    def process_semestral_sheet(self, sheet):
        """Enrutador hacia `imports.semestral_importer.process_semestral_sheet`."""
        return _process_semestral_sheet(sheet, self.audit_id)

    def process_auxiliary_records(self, sheet):
        """Enrutador hacia `imports.auxiliary_importer.process_auxiliary_records`."""
        return _process_auxiliary_records(sheet, self.audit_id)

    def process_initial_balances(self, sheet):
        """Enrutador hacia `imports.initial_balances_importer.process_initial_balances`."""
        return _process_initial_balances(sheet, self.audit_id)