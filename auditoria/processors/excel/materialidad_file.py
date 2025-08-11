from openpyxl.styles import Border, Side
from copy import copy

def process_materialidad_file(workbook, balances):
    """Procesa el archivo de materialidad con las cuentas de ESTADO DE RESULTADOS del año más reciente"""
    
    # Extraer todas las claves relacionadas con ESTADO DE RESULTADOS
    claves_resultados = [key for key in balances.keys() if 'ESTADO DE RESULTADOS' in key]
    
    # Extraer los años disponibles
    años = set()
    for key in claves_resultados:
        partes = key.split('-')
        if len(partes) >= 2:
            año = partes[1].split('-')[0]  # Extraer solo el año de la fecha
            años.add(año)
    
    if not años:
        return workbook
    
    # Encontrar el año más reciente
    año_reciente = max(años)
    
    # Filtrar las cuentas de ESTADO DE RESULTADOS del año más reciente
    cuentas_resultados = {}
    for key in claves_resultados:
        if año_reciente in key:
            partes = key.split('-')
            if len(partes) >= 6:  # Asegurarnos de que hay suficientes elementos
                cuenta = partes[5]  # La cuenta está en la posición 5, no en la 4
                valor = balances[key]
                cuentas_resultados[cuenta] = valor
    
    # Insertar las cuentas en la tabla de materialidad
    for sheet in workbook.worksheets:
        # Buscar la fila que contiene "Base de materialidad" usando el operador walrus
        if not (fila_base := next((idx for idx, row in enumerate(sheet.iter_rows(), 1) 
                                 if any(cell.value and isinstance(cell.value, str) 
                                        and "Base de materialidad" in cell.value for cell in row)), None)):
            continue
        
        # Insertar las cuentas en la tabla
        for i, (cuenta, valor) in enumerate(cuentas_resultados.items(), 1):
            # Obtener las celdas donde insertaremos los valores
            # Columna B (2) para el nombre de la cuenta, Columna D (4) para el valor
            celda_cuenta = sheet.cell(row=fila_base + i, column=2)
            celda_valor = sheet.cell(row=fila_base + i, column=4)
            
            # Insertar los valores
            celda_cuenta.value = cuenta
            celda_valor.value = valor
            
            # No copiamos ningún estilo, dejamos el estilo original de las celdas
    return workbook