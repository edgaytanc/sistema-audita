"""Operaciones sobre la tabla SUMARIA en la hoja de Excel.

Contiene:
• buscar_tabla_sumaria: detecta posiciones clave (filas/columnas).
• actualizar_fechas_encabezados: escribe las fechas semestrales en los
  encabezados estándar de la plantilla.
"""

from __future__ import annotations

from typing import Dict, List, Optional

from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet

__all__ = [
    "buscar_tabla_sumaria",
    "actualizar_fechas_encabezados",
]


# ---------------------------------------------------------------------
# Localizar tabla SUMARIA
# ---------------------------------------------------------------------

def buscar_tabla_sumaria(sheet: Worksheet) -> Optional[Dict[str, int | List[int]]]:
    """Intenta ubicar la tabla SUMARIA en *sheet*.

    Devuelve un diccionario con:
    - fila_titulo
    - fila_encabezados
    - fila_inicio_datos
    - columnas_fechas: lista de índices columna con leyenda "Saldos s/ Balance"
    - columna_cuenta: índice columna que contiene "Cuenta"/"Descripción"
    o **None** si la tabla no se encuentra.
    """

    fila_titulo = None
    for row_idx in range(1, 30):
        for col_idx in range(1, 15):
            val = sheet.cell(row=row_idx, column=col_idx).value
            if isinstance(val, str) and (
                "SUMARIA" in val.upper() or "CEDULA" in val.upper()
            ):
                fila_titulo = row_idx
                break
        if fila_titulo:
            break

    if not fila_titulo:
        return None

    fila_encabezados = None
    for row_idx in range(fila_titulo, fila_titulo + 15):
        for col_idx in range(1, 15):
            val = sheet.cell(row=row_idx, column=col_idx).value
            if isinstance(val, str) and "SALDOS S/ BALANCE" in val.upper():
                fila_encabezados = row_idx
                break
        if fila_encabezados:
            break

    if not fila_encabezados:
        return None

    columnas_fechas: List[int] = [
        col
        for col in range(1, 15)
        if isinstance(sheet.cell(row=fila_encabezados, column=col).value, str)
        and "SALDOS S/ BALANCE" in sheet.cell(row=fila_encabezados, column=col).value.upper()
    ]

    columna_cuenta = None
    for col_idx in range(1, 15):
        val = sheet.cell(row=fila_encabezados + 1, column=col_idx).value
        if isinstance(val, str) and (
            "CUENTA" in val.upper() or "DESCRIPCIÓN" in val.upper()
        ):
            columna_cuenta = col_idx
            break

    fila_inicio_datos = fila_encabezados + 2

    return {
        "fila_titulo": fila_titulo,
        "fila_encabezados": fila_encabezados,
        "fila_inicio_datos": fila_inicio_datos,
        "columnas_fechas": columnas_fechas,
        "columna_cuenta": columna_cuenta,
    }


# ---------------------------------------------------------------------
# Encabezados de fecha
# ---------------------------------------------------------------------

def _formatear_fecha_ddmmaa(fecha_iso: str) -> str:
    """Convierte 'YYYY-MM-DD' → 'DD/MM/YYYY'."""

    yyyy, mm, dd = fecha_iso.split("-")
    return f"{dd}/{mm}/{yyyy}"


def actualizar_fechas_encabezados(
    sheet: Worksheet,
    fechas_semestrales: List[str],
    tabla_info: Dict[str, int | List[int]],
) -> None:
    """Escribe las *fechas_semestrales* en las celdas estáticas de la plantilla."""

    fechas_ordenadas = sorted(fechas_semestrales)
    fechas_formateadas = [_formatear_fecha_ddmmaa(f) for f in fechas_ordenadas]

    celdas_destino = ["E12", "F12", "G12", "J12"]
    for idx, celda in enumerate(celdas_destino):
        if idx >= len(fechas_formateadas):
            break
        valor_actual = sheet[celda].value
        if isinstance(valor_actual, str) and valor_actual.upper().startswith("AL "):
            sheet[celda].value = f"AL {fechas_formateadas[idx]}"
        else:
            sheet[celda].value = fechas_formateadas[idx]
        sheet[celda].alignment = Alignment(horizontal="center")
