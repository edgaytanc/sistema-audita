"""Inserción de datos y manipulación de filas/estilos para la tabla SUMARIA."""

from __future__ import annotations

from typing import Dict, List

from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet

from .formulas import ajustar_formula_para_nueva_fila
from .data_extraction import encontrar_ajuste_para_cuenta
from .fechas import _formatear_fecha_ddmmaa

__all__ = [
    "insertar_datos_cuenta",
    "insertar_multiples_cuentas",
    "actualizar_fechas_encabezados",
]


# ---------------------------------------------------------------------
# Inserción sencilla (una cuenta)
# ---------------------------------------------------------------------

def insertar_datos_cuenta(
    sheet: Worksheet,
    datos_cuenta: Dict[str, float],
    tabla_info: Dict[str, int | List[int]],
    nombre_cuenta: str = "TOTAL",
    ajustes_reclasificaciones: Dict[str, Dict[str, float]] = None,
) -> None:
    """Escribe *datos_cuenta* en la fila 13 de la plantilla."""

    fechas_ordenadas = sorted(datos_cuenta.keys())
    sheet["B13"].value = nombre_cuenta
    celdas_valores = ["E13", "F13", "G13", "J13"]
    for idx, celda in enumerate(celdas_valores):
        if idx < len(fechas_ordenadas):
            fecha = fechas_ordenadas[idx]
            if fecha in datos_cuenta:
                sheet[celda].value = datos_cuenta[fecha]
                sheet[celda].alignment = Alignment(horizontal="right")

    # Insertar valores de ajustes/reclasificaciones usando la nueva función
    valores_ajuste = encontrar_ajuste_para_cuenta(ajustes_reclasificaciones, nombre_cuenta)
    if valores_ajuste:
        if 'debe' in valores_ajuste:
            sheet["H13"].value = valores_ajuste['debe']
            sheet["H13"].alignment = Alignment(horizontal="right")
        if 'haber' in valores_ajuste:
            sheet["I13"].value = valores_ajuste['haber']
            sheet["I13"].alignment = Alignment(horizontal="right")


# ---------------------------------------------------------------------
# Inserción múltiple (varias cuentas y filas dinámicas)
# ---------------------------------------------------------------------

def insertar_multiples_cuentas(
    sheet: Worksheet,
    todas_cuentas: Dict[str, Dict[str, float]],
    fechas_semestrales: List[str],
    tabla_info: Dict[str, int | List[int]],
    ajustes_reclasificaciones: Dict[str, Dict[str, float]] = None,
) -> None:
    """Inserta varias cuentas, creando filas adicionales después de la 13 si es necesario."""

    fechas_ordenadas = sorted(fechas_semestrales)
    cuentas_ordenadas = sorted(todas_cuentas.keys())

    filas_necesarias = len(cuentas_ordenadas)
    filas_disponibles = 1  # la plantilla tiene 1 fila vacía (13)
    filas_a_insertar = max(0, filas_necesarias - filas_disponibles)

    # Guardar rangos combinados de la fila 13
    celdas_combinadas_originales = []
    for merged_range in sheet.merged_cells.ranges:
        if 13 in range(merged_range.min_row, merged_range.max_row + 1):
            celdas_combinadas_originales.append(
                {
                    "min_row": merged_range.min_row,
                    "max_row": merged_range.max_row,
                    "min_col": merged_range.min_col,
                    "max_col": merged_range.max_col,
                }
            )

    if filas_a_insertar:
        # Copiar estilos/fórmulas de la fila 13
        estilos_fila_13 = {}
        formulas_fila_13 = {}
        for col in range(1, sheet.max_column + 1):
            celda = sheet.cell(row=13, column=col)
            estilos_fila_13[col] = {
                "font": celda.font.copy() if celda.font else None,
                "border": celda.border.copy() if celda.border else None,
                "fill": celda.fill.copy() if celda.fill else None,
                "number_format": celda.number_format,
                "alignment": celda.alignment.copy() if celda.alignment else None,
                "protection": celda.protection.copy() if celda.protection else None,
            }
            if celda.data_type == "f":
                formulas_fila_13[col] = celda.value

        sheet.insert_rows(14, filas_a_insertar)

        for i in range(filas_a_insertar):
            fila_destino = 14 + i
            for col, estilos in estilos_fila_13.items():
                celda_dest = sheet.cell(row=fila_destino, column=col)
                if estilos["font"]:
                    celda_dest.font = estilos["font"]
                if estilos["border"]:
                    celda_dest.border = estilos["border"]
                if estilos["fill"]:
                    celda_dest.fill = estilos["fill"]
                celda_dest.number_format = estilos["number_format"]
                if estilos["alignment"]:
                    celda_dest.alignment = estilos["alignment"]
                if estilos["protection"]:
                    celda_dest.protection = estilos["protection"]

                # Ajustar fórmulas
                if col in formulas_fila_13:
                    formula_orig = formulas_fila_13[col]
                    if isinstance(formula_orig, str) and formula_orig.startswith("="):
                        celda_dest.value = ajustar_formula_para_nueva_fila(
                            formula_orig, 13, fila_destino
                        )

            # Restaurar combinaciones
            for rng in celdas_combinadas_originales:
                if rng["max_col"] > rng["min_col"]:
                    rango = f"{get_column_letter(rng['min_col'])}{fila_destino}:{get_column_letter(rng['max_col'])}{fila_destino}"
                    sheet.merge_cells(rango)

    # Escribir datos en filas
    for i, nombre_cuenta in enumerate(cuentas_ordenadas):
        fila = 13 + i
        sheet[f"B{fila}"].value = nombre_cuenta
        celdas_valores_col = ["E", "F", "G", "J"]
        for j, col_letra in enumerate(celdas_valores_col):
            if j < len(fechas_ordenadas):
                fecha = fechas_ordenadas[j]
                if fecha in todas_cuentas[nombre_cuenta]:
                    sheet[f"{col_letra}{fila}"].value = todas_cuentas[nombre_cuenta][fecha]
                    sheet[f"{col_letra}{fila}"].alignment = Alignment(horizontal="right")

        # Insertar valores de ajustes/reclasificaciones usando la nueva función
        valores_ajuste = encontrar_ajuste_para_cuenta(ajustes_reclasificaciones, nombre_cuenta)
        if valores_ajuste:
            if 'debe' in valores_ajuste:
                sheet[f"H{fila}"].value = valores_ajuste['debe']
                sheet[f"H{fila}"].alignment = Alignment(horizontal="right")
            if 'haber' in valores_ajuste:
                sheet[f"I{fila}"].value = valores_ajuste['haber']
                sheet[f"I{fila}"].alignment = Alignment(horizontal="right")


# ---------------------------------------------------------------------
# Actualización de encabezados de fecha
# ---------------------------------------------------------------------

def actualizar_fechas_encabezados(
    sheet: Worksheet,
    fechas_semestrales: List[str],
    tabla_info: Dict[str, int | List[int]],
) -> None:
    """Escribe las *fechas_semestrales* en las celdas estáticas de la plantilla."""

    fechas_ordenadas = sorted(fechas_semestrales)
    fechas_formateadas = [_formatear_fecha_ddmmaa(f) for f in fechas_ordenadas]

    celdas_destino = ["E12", "F12", "G12", "J12"]
    for idx, celda in enumerate(celdas_destino):
        if idx >= len(fechas_formateadas):
            break
        valor_actual = sheet[celda].value
        if isinstance(valor_actual, str) and valor_actual.upper().startswith("AL "):
            sheet[celda].value = f"AL {fechas_formateadas[idx]}"
        else:
            sheet[celda].value = fechas_formateadas[idx]
        sheet[celda].alignment = Alignment(horizontal="center")
