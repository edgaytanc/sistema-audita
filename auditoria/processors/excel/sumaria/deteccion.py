"""Detección de cuenta asociada y verificación en balances para procesadores SUMARIA."""

from __future__ import annotations

import re
from typing import Tuple, Dict, List, Optional

from openpyxl.worksheet.worksheet import Worksheet

__all__ = [
    "determinar_cuenta_por_nombre_archivo",
    "verificar_cuenta_en_balances",
    "buscar_tabla_sumaria",
]


# ---------------------------------------------------------------------
# Mapeo de palabras clave a nombres de cuentas contables
# ---------------------------------------------------------------------

_MAPEO_CUENTAS: dict[str, str] = {
    "CAJA": "Caja y Bancos",
    "BANCO": "Caja y Bancos",
    "COBRAR": "Cuentas por Cobrar",
    "CLIENTE": "Cuentas por Cobrar",
    "INVENTARIO": "Inventarios",
    "INVERSION": "Inversiones",
    "CONSTRUCCION": "Construcciones en Proceso",
    "PROCESO": "Construcciones en Proceso",
    "ACTIVO FIJO": "Propiedad, Planta y Equipo",
    "PROPIEDAD": "Propiedad, Planta y Equipo",
    "PLANTA": "Propiedad, Planta y Equipo",
    "EQUIPO": "Propiedad, Planta y Equipo",
    "INTANGIBLE": "Activo Intangible",
    "PAGAR": "Cuentas por Pagar",
    "PROVEEDOR": "Cuentas por Pagar",
    "PASIVO LARGO PLAZO": "Pasivo Largo Plazo",
    "PASIVO": "Pasivo",
    "PRESTAMO": "Préstamos por Pagar",
    "FINANCIAMIENTO": "Préstamos por Pagar",
    "PATRIMONIO": "Patrimonio",
    "CAPITAL": "Capital",
    "RESULTADO": "Estado de Resultados",
    "INGRESO": "Ingresos",
    "EGRESO": "Egresos",
    "GASTO": "Gastos",
    "COSTO": "Costos",
}


# ---------------------------------------------------------------------
# Detección de cuenta por nombre de archivo
# ---------------------------------------------------------------------

def determinar_cuenta_por_nombre_archivo(file_name: str) -> str | None:
    """Intenta inferir la cuenta contable a partir del nombre del archivo.

    Devuelve el nombre normalizado de la cuenta o None si no puede determinarse.
    """

    # Quitar extensión y números iniciales
    nombre_limpio = re.sub(r"^\d+\s*", "", file_name.replace(".xlsx", ""))

    # Buscar por palabras clave, de más largas a más cortas
    for palabra_clave in sorted(_MAPEO_CUENTAS, key=len, reverse=True):
        if palabra_clave in nombre_limpio.upper():
            return _MAPEO_CUENTAS[palabra_clave]

    # Fallback: extraer texto posterior a "SUMARIA "
    if m := re.search(r"SUMARIA\s+(.*?)(?:\s+|$)", nombre_limpio.upper()):
        return m.group(1).title()

    return None


# ---------------------------------------------------------------------
# Verificación de existencia de cuenta en balances
# ---------------------------------------------------------------------

def verificar_cuenta_en_balances(
    balances: Dict[str, float],
    cuenta_asociada: str,
    fechas_semestrales: List[str],
) -> Tuple[bool, Dict[str, float]]:
    """Comprueba la existencia de *cuenta_asociada* en *balances* y
    devuelve los valores por fecha.

    Retorna `(existe, datos_por_fecha)`.
    """

    datos_cuenta: dict[str, float] = {}
    cuenta_existe = False
    cuenta_normalizada = cuenta_asociada.lower().strip()

    for seccion in ["Activo", "Pasivo", "Patrimonio", "ESTADO DE RESULTADOS"]:
        for fecha in fechas_semestrales:
            clave_exacta = f"SEMESTRAL-{fecha}-{seccion}-{cuenta_asociada}"
            if clave_exacta in balances:
                datos_cuenta[fecha] = balances[clave_exacta]
                cuenta_existe = True
                continue

            # Buscar coincidencias parciales si no hay exactas
            if fecha not in datos_cuenta:
                prefijo = f"SEMESTRAL-{fecha}-{seccion}-"
                for clave, valor in balances.items():
                    if clave.startswith(prefijo):
                        nombre_cuenta = "-".join(clave.split("-")[3:]).lower().strip()
                        if cuenta_normalizada in nombre_cuenta or nombre_cuenta in cuenta_normalizada:
                            datos_cuenta[fecha] = valor
                            cuenta_existe = True
                            break

    return cuenta_existe, datos_cuenta


# ---------------------------------------------------------------------
# Localización de tabla SUMARIA en hoja Excel
# ---------------------------------------------------------------------

def buscar_tabla_sumaria(sheet: Worksheet) -> Optional[Dict[str, int | List[int]]]:
    """Intenta ubicar la tabla SUMARIA en *sheet*.

    Devuelve un diccionario con:
    - fila_titulo
    - fila_encabezados
    - fila_inicio_datos
    - columnas_fechas: lista de índices columna con leyenda "Saldos s/ Balance"
    - columna_cuenta: índice columna que contiene "Cuenta"/"Descripción"
    o **None** si la tabla no se encuentra.
    """

    fila_titulo = None
    for row_idx in range(1, 30):
        for col_idx in range(1, 15):
            val = sheet.cell(row=row_idx, column=col_idx).value
            if isinstance(val, str) and (
                "SUMARIA" in val.upper() or "CEDULA" in val.upper()
            ):
                fila_titulo = row_idx
                break
        if fila_titulo:
            break

    if not fila_titulo:
        return None

    fila_encabezados = None
    for row_idx in range(fila_titulo, fila_titulo + 15):
        for col_idx in range(1, 15):
            val = sheet.cell(row=row_idx, column=col_idx).value
            if isinstance(val, str) and "SALDOS S/ BALANCE" in val.upper():
                fila_encabezados = row_idx
                break
        if fila_encabezados:
            break

    if not fila_encabezados:
        return None

    columnas_fechas: List[int] = [
        col
        for col in range(1, 15)
        if isinstance(sheet.cell(row=fila_encabezados, column=col).value, str)
        and "SALDOS S/ BALANCE" in sheet.cell(row=fila_encabezados, column=col).value.upper()
    ]

    columna_cuenta = None
    for col_idx in range(1, 15):
        val = sheet.cell(row=fila_encabezados + 1, column=col_idx).value
        if isinstance(val, str) and (
            "CUENTA" in val.upper() or "DESCRIPCIÓN" in val.upper()
        ):
            columna_cuenta = col_idx
            break

    fila_inicio_datos = fila_encabezados + 2

    return {
        "fila_titulo": fila_titulo,
        "fila_encabezados": fila_encabezados,
        "fila_inicio_datos": fila_inicio_datos,
        "columnas_fechas": columnas_fechas,
        "columna_cuenta": columna_cuenta,
    }
