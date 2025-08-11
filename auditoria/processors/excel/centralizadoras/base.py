import re
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment



def obtener_todas_fechas_semestrales(balances):
    """
    Obtiene todas las fechas semestrales disponibles en los balances

    Args:
        balances: Diccionario con los balances financieros

    Returns:
        Lista con todas las fechas semestrales en formato YYYY-MM-DD
    """
    patron_fecha = re.compile(r'SEMESTRAL-(\d{4}-\d{2}-\d{2})')
    fechas = set()

    for clave in balances:
        match = patron_fecha.search(clave)
        if match:
            fecha_completa = match.group(1)
            fechas.add(fecha_completa)

    # Devolver las fechas ordenadas
    return sorted(list(fechas))


def filtrar_cuentas_por_seccion(balances, fecha, seccion):
    """
    Filtra las cuentas de una sección específica para una fecha dada

    Args:
        balances: Diccionario con los balances financieros
        fecha: Fecha en formato YYYY-MM-DD
        seccion: Nombre de la sección (Activo, Pasivo, Patrimonio, ESTADO DE RESULTADOS)

    Returns:
        Diccionario con las cuentas y sus valores
    """
    cuentas_filtradas = {}
    patron = re.compile(f'SEMESTRAL-{fecha}-{seccion}-(.*)')

    for clave, valor in balances.items():
        match = patron.match(clave)
        if match:
            nombre_cuenta = match.group(1)
            cuentas_filtradas[nombre_cuenta] = valor

    return cuentas_filtradas


def preparar_fechas_excel(fechas_semestrales):
    """
    Prepara las fechas para mostrar en Excel

    Args:
        fechas_semestrales: Lista con todas las fechas semestrales

    Returns:
        Tuple con (fechas_año_viejo, fecha_mas_reciente, año_mas_viejo, año_mas_reciente, fechas_excel)
    """
    # Ordenar fechas por año
    fechas_por_año = {}
    for fecha in fechas_semestrales:
        año = fecha.split('-')[0]
        if año not in fechas_por_año:
            fechas_por_año[año] = []
        fechas_por_año[año].append(fecha)

    # Ordenar años
    años_ordenados = sorted(fechas_por_año.keys())

    if not años_ordenados:
        return None, None, None, None, None

    # Obtener el año más viejo y el más reciente
    año_mas_viejo = años_ordenados[0]
    año_mas_reciente = años_ordenados[-1]

    # Obtener las fechas del año más viejo (ordenadas)
    fechas_año_viejo = sorted(fechas_por_año[año_mas_viejo])

    # Verificar si tenemos suficientes fechas para el año más viejo
    while len(fechas_año_viejo) < 3:
        if fechas_año_viejo:
            fechas_año_viejo.append(fechas_año_viejo[-1])
        else:
            fechas_año_viejo.append("")

    # Obtener la fecha más reciente del año más reciente
    fecha_mas_reciente = sorted(fechas_por_año[año_mas_reciente])[-1]

    # Formatear fechas para Excel
    fechas_excel = {
        'D12': f"Al 01/01/{año_mas_viejo}",
        'E12': f"Al 31/07/{año_mas_viejo}",
        'F12': f"Al 31/12/{año_mas_viejo}",
        'I12': f"Al 31/12/{año_mas_reciente}"
    }

    return fechas_año_viejo, fecha_mas_reciente, año_mas_viejo, año_mas_reciente, fechas_excel


def insertar_fechas_en_celdas(sheet, fechas_excel):
    """
    Inserta las fechas en las celdas correspondientes

    Args:
        sheet: Hoja de Excel
        fechas_excel: Diccionario con las fechas a insertar
    """
    for celda, fecha in fechas_excel.items():
        sheet[celda] = fecha
        sheet[celda].alignment = Alignment(horizontal='center')


def procesar_seccion_balance(
    sheet,
    cuentas_por_fecha,
    fechas_año_viejo,
    fecha_mas_reciente,
    seccion,
    fila_inicio,
    fila_fin,
    orden_explicit: list | None = None,
):
    """
    Procesa una sección específica del balance (Activo, Pasivo, Patrimonio)

    Args:
        sheet: Hoja de Excel a procesar
        cuentas_por_fecha: Diccionario con las cuentas por fecha
        fechas_año_viejo: Lista con las fechas del año más viejo
        fecha_mas_reciente: Fecha más reciente
        seccion: Nombre de la sección (Activo, Pasivo, Patrimonio)
        fila_inicio: Fila donde comienza la sección
        fila_fin: Fila donde termina la sección
        orden_explicit: Lista de cuentas en orden explícito (opcional)
    """
    # Obtener los estilos de las celdas de la fila de inicio para copiarlos
    estilos_fila = {}
    for col in range(1, 10):  # Columnas A a I
        col_letra = get_column_letter(col)
        celda = sheet[f'{col_letra}{fila_inicio}']
        estilos_fila[col_letra] = {
            'font': celda.font.copy() if celda.font else None,
            'border': celda.border.copy() if celda.border else None,
            'fill': celda.fill.copy() if celda.fill else None,
            'number_format': celda.number_format,
            'alignment': celda.alignment.copy() if celda.alignment else None,
            'protection': celda.protection.copy() if celda.protection else None
        }

    # Obtener todas las cuentas para esta sección
    todas_las_cuentas = set()
    for fecha in fechas_año_viejo + [fecha_mas_reciente]:
        if fecha in cuentas_por_fecha:
            # Asegurarnos de que estamos accediendo correctamente a las cuentas según la sección
            if seccion in cuentas_por_fecha[fecha]:
                todas_las_cuentas.update(
                    cuentas_por_fecha[fecha][seccion].keys())

    cuentas_ordenadas = orden_explicit if orden_explicit is not None else sorted(
        todas_las_cuentas)

    # Calcular cuántas filas necesitamos
    filas_disponibles = fila_fin - fila_inicio + 1
    filas_necesarias = len(cuentas_ordenadas)

    # Si necesitamos más filas, insertarlas
    if filas_necesarias > filas_disponibles:
        filas_a_insertar = filas_necesarias - filas_disponibles

        sheet.insert_rows(fila_fin + 1, filas_a_insertar)

        # Copiar los estilos a las nuevas filas
        for fila in range(fila_fin + 1, fila_fin + 1 + filas_a_insertar):
            for col_letra, estilos in estilos_fila.items():
                celda_destino = sheet[f'{col_letra}{fila}']

                if estilos['font']:
                    celda_destino.font = estilos['font']
                if estilos['border']:
                    celda_destino.border = estilos['border']
                if estilos['fill']:
                    celda_destino.fill = estilos['fill']
                celda_destino.number_format = estilos['number_format']
                if estilos['alignment']:
                    celda_destino.alignment = estilos['alignment']
                if estilos['protection']:
                    celda_destino.protection = estilos['protection']

    # Insertar los datos en las filas
    for i, cuenta in enumerate(cuentas_ordenadas):
        fila_actual = fila_inicio + i

        # Insertar el nombre de la cuenta en la columna B
        sheet[f'B{fila_actual}'] = cuenta

        # Insertar valores para las diferentes fechas
        valores = {
            'D': fechas_año_viejo[0],
            'E': fechas_año_viejo[1],
            'F': fechas_año_viejo[2],
            'I': fecha_mas_reciente
        }

        for col, fecha in valores.items():
            if fecha in cuentas_por_fecha and seccion in cuentas_por_fecha[fecha]:
                if cuenta in cuentas_por_fecha[fecha][seccion]:
                    sheet[f'{col}{fila_actual}'] = cuentas_por_fecha[fecha][seccion][cuenta]
                else:
                    sheet[f'{col}{fila_actual}'] = 0
            else:
                sheet[f'{col}{fila_actual}'] = 0


def buscar_fila_por_valor(sheet, col_letter: str, texto: str):
    """Devuelve el número de fila donde la celda de la columna col_letter contiene 'texto'."""
    texto_lower = texto.lower()
    for row in range(1, sheet.max_row + 1):
        val = sheet[f"{col_letter}{row}"].value
        if isinstance(val, str) and texto_lower in val.lower():
            return row
    return None
